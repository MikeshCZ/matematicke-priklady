#!/usr/bin/env python3
"""
Generator matematickych prikladu do Excel souboru.
Podporuje scitani, odcitani, nasobeni a deleni s konfigurovatelnym max vysledkem.
"""
import random
import argparse
import math
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


# ----------------------------
# Generovani prikladu
# ----------------------------
def gen_add(max_result):
    """
    Generuje priklad na scitani.

    Args:
        max_result: Maximalni vysledek operace

    Returns:
        Tuple (a, "+", b, vysledek) kde a + b <= max_result
    """
    a = random.randint(0, max_result)
    b = random.randint(0, max_result - a)
    return a, "+", b, a + b


def gen_sub(max_result):
    """
    Generuje priklad na odcitani.

    Args:
        max_result: Maximalni hodnota mensence

    Returns:
        Tuple (a, "-", b, vysledek) kde a >= b (nezaporne vysledky)
    """
    a = random.randint(0, max_result)
    b = random.randint(0, a)
    return a, "-", b, a - b


def gen_mul(max_result):
    """
    Generuje priklad na nasobeni.

    Args:
        max_result: Maximalni vysledek operace

    Returns:
        Tuple (a, "*", b, vysledek) kde a * b <= max_result

    Note:
        Pouziva kandidatsky pristup pro rovnomernou distribuci operandu
    """
    candidates = []
    for a in range(0, max_result + 1):
        if a == 0:
            candidates.append((0, "*", 0, 0))
            continue
        max_b = max_result // a
        b = random.randint(0, max_b)
        candidates.append((a, "*", b, a * b))
    return random.choice(candidates)


def gen_div(max_result):
    """
    Generuje priklad na deleni.

    Args:
        max_result: Maximalni hodnota vysledku

    Returns:
        Tuple (a, "/", b, vysledek) kde a / b = cely vysledek bez zbytku

    Note:
        Vysledek je vzdy cele cislo (a = b * vysledek)
    """
    if max_result <= 0:
        return 0, "/", 1, 0
    c = random.randint(0, max_result)
    b = random.randint(1, max(1, max_result))
    a = b * c
    return a, "/", b, c


# Mapovani operacnich symbolu na generatory
# Podporuje aliasy: 'x' pro '*' a '÷' pro '/'
GEN_MAP = {
    "+": gen_add,
    "-": gen_sub,
    "*": gen_mul,
    "x": gen_mul,
    "/": gen_div,
    "÷": gen_div,
}


def make_problem_text(max_result, ops):
    """
    Vytvori textovou reprezentaci prikladu.

    Args:
        max_result: Maximalni vysledek pro generovani
        ops: Seznam operaci k pouziti

    Returns:
        String ve formatu "a op b = ___"
    """
    op_key = random.choice(ops)
    a, op_sym, b, _ = GEN_MAP[op_key](max_result)
    return f"{a} {op_sym} {b} = ___"


# ----------------------------
# Generovani Excel listu
# ----------------------------
def generate_sheet(
    ops, max_result, count, file_name, seed=None, title=None, cols=2, fill_mode="down"
):
    """
    Hlavni funkce pro generovani Excel souboru s priklady.

    Args:
        ops: Seznam operaci k pouziti ('+', '-', '*', '/')
        max_result: Maximalni vysledek prikladu
        count: Pocet prikladu k vygenerovani
        file_name: Nazev vystupniho .xlsx souboru
        seed: Volitelny seed pro reprodukovatelnost (default: None)
        title: Volitelny titulek listu (default: None)
        cols: Pocet sloupcu v rozlozeni (default: 2)
        fill_mode: Zpusob vyplnovani "down" (po sloupcich) nebo "across" (po radcich)

    Returns:
        Cesta k vytvorenememu souboru

    Raises:
        ValueError: Pokud nejsou zadany platne operace
    """
    # Nastaveni seedu pro reprodukovatelnost
    if seed is not None:
        random.seed(seed)

    # Filtrace platnych operaci
    ops = [o for o in ops if o in GEN_MAP]
    if not ops:
        raise ValueError("Zadna platna operace (+ - * /).")

    # Validace poctu sloupcu
    cols = max(1, cols)

    # Vytvoreni Excel workbooku
    wb = Workbook()
    ws = wb.active
    ws.title = "Priklady"

    # Stylovani - pouzijeme monospace font pro spravne zarovnani
    font = Font(name="Consolas", size=16)
    align_left = Alignment(horizontal="left", vertical="center")

    # Hlavicka / titulek
    start_row = 1
    if title:
        title_cell = ws.cell(row=start_row, column=1)
        title_cell.value = title
        title_cell.font = Font(name="Calibri", size=18, bold=True)
        start_row += 2

    # Vygenerovani vsech prikladu najednou
    problems = [make_problem_text(max_result, ops) for _ in range(count)]

    # Validace fill_mode
    if fill_mode not in ("down", "across"):
        fill_mode = "down"

    # Vypocet potrebneho poctu radku
    rows_needed = math.ceil(count / cols)

    # Rozdeleni prikladu do sloupcu pro zjisteni max delky v kazdem sloupci
    # Nejprve vytvorime 2D strukturu prikladu podle fill_mode
    grid = [[None for _ in range(cols)] for _ in range(rows_needed)]
    idx = 0

    if fill_mode == "down":
        # Column-major: plneni po sloupcich
        for c in range(cols):
            for r in range(rows_needed):
                if idx >= count:
                    break
                grid[r][c] = problems[idx]
                idx += 1
    else:  # across
        # Row-major: plneni po radcich
        for r in range(rows_needed):
            for c in range(cols):
                if idx >= count:
                    break
                grid[r][c] = problems[idx]
                idx += 1

    # Najdeme maximalni delku leve casti (pred "= ___") v kazdem sloupci
    max_lens = [0] * cols
    for c in range(cols):
        for r in range(rows_needed):
            if grid[r][c] is not None:
                # Ziskame cast pred "= ___"
                left_part = grid[r][c].split(" = ")[0]
                max_lens[c] = max(max_lens[c], len(left_part))

    # Zarovname priklady v kazdem sloupci podle nejdelsiho
    # Mezery budou na zacatku pro zarovnani k "="
    for c in range(cols):
        for r in range(rows_needed):
            if grid[r][c] is not None:
                parts = grid[r][c].split(" = ")
                left_part = parts[0]
                # Pridame mezery na zacatek pro zarovnani zprava
                padding = max_lens[c] - len(left_part)
                grid[r][c] = " " * padding + left_part + " = ___"

    # Zapsani zarovnanych prikladu do buniek
    for r in range(rows_needed):
        for c in range(cols):
            if grid[r][c] is not None:
                cell = ws.cell(row=start_row + r, column=1 + c)
                cell.value = grid[r][c]
                cell.font = font
                cell.alignment = align_left

    # Nastaveni jednotne sirky sloupcu (cca 200 px)
    EXCEL_WIDTH = 29
    for c in range(1, cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = EXCEL_WIDTH

    # Zvyseni radkovani pro lepsi citelnost
    ROW_HEIGHT = 24
    for r in range(start_row, start_row + rows_needed):
        ws.row_dimensions[r].height = ROW_HEIGHT

    # Nastaveni uzkych okraju (narrow margins) pro tisk
    # Hodnoty jsou v palcich
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3

    # Ulozeni souboru
    wb.save(file_name)
    return file_name


# ----------------------------
# CLI rozhrani
# ----------------------------
def parse_args():
    """
    Parsuje argumenty prikazove radky.

    Returns:
        Namespace s argumenty CLI
    """
    p = argparse.ArgumentParser(
        prog="python src/main.py",
        description="Generator matematickych prikladu do Excel souboru (.xlsx)",
        epilog="Pro GUI rozhrani pouzijte: python src/gui.py",
        formatter_class=argparse.RawDescriptionHelpFormatter
    )

    p.add_argument(
        "--ops",
        type=str,
        default="+-*/",
        metavar="OPERACE",
        help="Operace k pouziti. Retezec obsahujici '+' (scitani), '-' (odcitani), '*' (nasobeni), '/' (deleni). Vychozi: '+-*/'",
    )
    p.add_argument(
        "--max",
        type=int,
        default=20,
        metavar="CISLO",
        help="Maximalni vysledek prikladu. Rozsah: 1-1000. Vychozi: 20",
    )
    p.add_argument(
        "--count",
        type=int,
        default=78,
        metavar="CISLO",
        help="Pocet prikladu k vygenerovani. Rozsah: 1-500. Vychozi: 78",
    )
    p.add_argument(
        "--seed",
        type=int,
        default=None,
        metavar="CISLO",
        help="Seed pro reprodukovatelne generovani. Pouzijte stejne cislo pro identicky vysledek. Vychozi: nahodny",
    )
    p.add_argument(
        "--out",
        type=str,
        default="priklady.xlsx",
        metavar="SOUBOR",
        help="Nazev vystupniho .xlsx souboru. Vychozi: 'priklady.xlsx'",
    )
    p.add_argument(
        "--title",
        type=str,
        default="Matematicke priklady",
        metavar="TEXT",
        help="Titulek zobrazeny v hlavicce listu. Vychozi: 'Matematicke priklady'",
    )
    p.add_argument(
        "--cols",
        type=int,
        default=3,
        metavar="CISLO",
        help="Pocet sloupcu v rozlozeni. Rozsah: 1-10. Vychozi: 3",
    )
    p.add_argument(
        "--fill",
        type=str,
        default="down",
        choices=["down", "across"],
        metavar="REZIM",
        help="Zpusob vyplnovani: 'down' (po sloupcich shora dolu) nebo 'across' (po radcich zleva doprava). Vychozi: 'down'",
    )

    return p.parse_args()


def main():
    """Hlavni entry point pro CLI aplikaci."""
    args = parse_args()
    file_path = generate_sheet(
        ops=list(args.ops),
        max_result=args.max,
        count=args.count,
        file_name=args.out,
        seed=args.seed,
        title=args.title,
        cols=args.cols,
        fill_mode=args.fill,
    )
    print(f"Hotovo: {file_path}")


if __name__ == "__main__":
    main()
