#!/usr/bin/env python3
"""
Generator matematickych prikladu do Excel souboru.
Podporuje scitani, odcitani, nasobeni a deleni s konfigurovatelnym max vysledkem.
"""
import random
import argparse
import math
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

__version__ = "1.1.0"


# ----------------------------
# Generovani prikladu
# ----------------------------
def gen_add(max_result=None, max_digits=None, no_zero=False):
    """
    Generuje priklad na scitani.

    Args:
        max_result: Maximalni vysledek operace (omezuje vysledek a+b)
        max_digits: Maximalni pocet cislic v cislech (omezuje a a b)
        no_zero: Pokud True, vyloucit nulu z cisel (default: False)

    Returns:
        Tuple (a, "+", b, vysledek) kde a + b <= max_result

    Note:
        Pokud jsou zadany oba parametry, pouzije se prisnejsi limit.
        max_digits omezuje jednotliva cisla, max_result omezuje vysledek.
    """
    # Urceni maximalni hodnoty pro jednotliva cisla
    max_number = None
    if max_digits is not None:
        max_number = 10 ** max_digits - 1

    # Pokud neni zadano ani jedno, pouzijeme vychozi 2 cislice
    if max_result is None and max_number is None:
        max_result = 99
        max_number = 99
    elif max_number is None:
        max_number = max_result
    elif max_result is None:
        max_result = max_number * 2  # soucet dvou max cisel

    min_val = 1 if no_zero else 0
    a = random.randint(min_val, min(max_number, max_result))
    max_b_result = max_result - a
    max_b = min(max_number, max_b_result)
    # Zajistit, ze max_b je alespon min_val
    if max_b < min_val:
        # Pokud max_b je moc male, zkusime mensi a
        a = random.randint(min_val, min(max_number, max_result - min_val))
        max_b = min(max_number, max_result - a)
    b = random.randint(min_val, max_b)
    return a, "+", b, a + b


def gen_sub(max_result=None, max_digits=None, no_zero=False):
    """
    Generuje priklad na odcitani.

    Args:
        max_result: Maximalni hodnota mensence (omezuje a)
        max_digits: Maximalni pocet cislic v cislech (omezuje a a b)
        no_zero: Pokud True, vyloucit nulu z cisel a vysledku (default: False)

    Returns:
        Tuple (a, "-", b, vysledek) kde a >= b (nezaporne vysledky)

    Note:
        Pokud jsou zadany oba parametry, pouzije se prisnejsi limit.
    """
    # Urceni maximalni hodnoty pro jednotliva cisla
    max_number = None
    if max_digits is not None:
        max_number = 10 ** max_digits - 1

    # Pokud neni zadano ani jedno, pouzijeme vychozi 2 cislice
    if max_result is None and max_number is None:
        max_result = 99
        max_number = 99
    elif max_number is None:
        max_number = max_result
    elif max_result is None:
        max_result = max_number

    min_val = 1 if no_zero else 0
    # Pokud no_zero, a musi byt alespon 2, aby b mohlo byt alespon 1 a vysledek != 0
    min_a = 2 if no_zero else min_val
    a = random.randint(min_a, min(max_number, max_result))
    # Pokud no_zero, b musi byt alespon o 1 mensi nez a, aby vysledek nebyl 0
    max_b = a - 1 if no_zero else a
    b = random.randint(min_val, min(max_number, max_b))
    return a, "-", b, a - b


def gen_mul(max_result=None, max_digits=None, no_zero=False):
    """
    Generuje priklad na nasobeni.

    Args:
        max_result: Maximalni vysledek operace (omezuje vysledek a*b)
        max_digits: Maximalni pocet cislic v cislech (omezuje a a b)
        no_zero: Pokud True, vyloucit nulu z cisel (default: False)

    Returns:
        Tuple (a, "*", b, vysledek) kde a * b <= max_result

    Note:
        Pouziva kandidatsky pristup pro rovnomernou distribuci operandu.
        Pokud jsou zadany oba parametry, pouzije se prisnejsi limit.
    """
    # Urceni maximalni hodnoty pro jednotliva cisla
    max_number = None
    if max_digits is not None:
        max_number = 10 ** max_digits - 1

    # Pokud neni zadano ani jedno, pouzijeme vychozi 2 cislice
    if max_result is None and max_number is None:
        max_result = 99
        max_number = 99
    elif max_number is None:
        max_number = max_result
    elif max_result is None:
        max_result = max_number * max_number  # soucin dvou max cisel

    min_val = 1 if no_zero else 0
    candidates = []
    max_a = min(max_number, max_result)
    # Zajistit, ze max_a je alespon min_val
    if max_a < min_val:
        max_a = min_val

    for a in range(min_val, max_a + 1):
        if a == 0 and not no_zero:
            candidates.append((0, "×", 0, 0))
            continue
        if a == 0:  # Skip 0 when no_zero
            continue
        max_b = max_result // a
        max_b = min(max_b, max_number)  # omezit i b podle max_number
        if max_b < min_val:
            continue
        b = random.randint(min_val, max_b)
        candidates.append((a, "×", b, a * b))

    # Fallback pokud nejsou zadne candidates
    if not candidates:
        return (min_val, "×", min_val, min_val * min_val)
    return random.choice(candidates)


def gen_div(max_result=None, max_digits=None, no_zero=False):
    """
    Generuje priklad na deleni.

    Args:
        max_result: Maximalni hodnota vysledku (omezuje vysledek a/b)
        max_digits: Maximalni pocet cislic v cislech (omezuje a a b)
        no_zero: Pokud True, vyloucit nulu z cisel (default: False)

    Returns:
        Tuple (a, "/", b, vysledek) kde a / b = cely vysledek bez zbytku

    Note:
        Vysledek je vzdy cele cislo (a = b * vysledek).
        Pokud jsou zadany oba parametry, pouzije se prisnejsi limit.
    """
    # Urceni maximalni hodnoty pro jednotliva cisla
    max_number = None
    if max_digits is not None:
        max_number = 10 ** max_digits - 1

    # Pokud neni zadano ani jedno, pouzijeme vychozi 2 cislice
    if max_result is None and max_number is None:
        max_result = 99
        max_number = 99
    elif max_number is None:
        max_number = max_result * max_result  # deleni muze mit velke delence
    elif max_result is None:
        max_result = max_number

    if max_result <= 0:
        return 1, "/", 1, 1 if no_zero else (0, "/", 1, 0)

    min_val = 1 if no_zero else 0
    c = random.randint(min_val, min(max_number, max_result))  # vysledek
    b = random.randint(1, min(max_number, max(1, max_result)))  # delitel (nikdy nula)
    a = b * c  # delenec

    # Kontrola: pokud a prekracuje max_number nebo max_result, musime upravit b nebo c
    # Chceme zajistit, ze VSECHNA cisla (a, b, c) jsou v limitech
    max_a = min(max_number, max_result) if max_result is not None else max_number
    if a > max_a:
        # Zkusime najit validni kombinaci b a c
        # a = b * c, tedy b <= max_a / c
        max_b_allowed = max_a // c if c > 0 else max_a
        if max_b_allowed < 1:
            # c je prilis velke, zkusime mensi c
            c = random.randint(min_val, min(max_number, max_result, max_a // 2))
            max_b_allowed = max_a // c if c > 0 else 1
        b = random.randint(1, min(max_number, max_b_allowed, max_result))
        a = b * c

    return a, "/", b, c


# Mapovani operacnich symbolu na generatory
# Podporuje aliasy: 'x' pro '*' a '÷' pro '/'
GEN_MAP = {
    "+": gen_add,
    "-": gen_sub,
    "*": gen_mul,
    "x": gen_mul,
    "/": gen_div,
    "÷": gen_div,
}


def make_problem_text(ops, max_result=None, max_digits=None, no_zero=False):
    """
    Vytvori textovou reprezentaci prikladu.

    Args:
        ops: Seznam operaci k pouziti
        max_result: Maximalni vysledek pro generovani (deprecated)
        max_digits: Maximalni pocet cislic v cislech (doporuceno)
        no_zero: Pokud True, vyloucit nulu z cisel (default: False)

    Returns:
        String ve formatu "a op b = ___"
    """
    op_key = random.choice(ops)
    a, op_sym, b, _ = GEN_MAP[op_key](max_result=max_result, max_digits=max_digits, no_zero=no_zero)
    return f"{a} {op_sym} {b} = ___"


# ----------------------------
# Generovani Excel listu
# ----------------------------
def generate_sheet(
    ops, count, file_name, max_result=None, max_digits=None, seed=None, title=None, cols=2, fill_mode="down", no_zero=False
):
    """
    Hlavni funkce pro generovani Excel souboru s priklady.

    Args:
        ops: Seznam operaci k pouziti ('+', '-', '*', '/')
        count: Pocet prikladu k vygenerovani
        file_name: Nazev vystupniho .xlsx souboru
        max_result: Maximalni vysledek prikladu (deprecated, pouzijte max_digits)
        max_digits: Maximalni pocet cislic v cislech (doporuceno, vychozi: 2)
        seed: Volitelny seed pro reprodukovatelnost (default: None)
        title: Volitelny titulek listu (default: None)
        cols: Pocet sloupcu v rozlozeni (default: 2)
        fill_mode: Zpusob vyplnovani "down" (po sloupcich) nebo "across" (po radcich)
        no_zero: Pokud True, vyloucit nulu z cisel (default: False)

    Returns:
        Cesta k vytvorenememu souboru

    Raises:
        ValueError: Pokud nejsou zadany platne operace
    """
    # Nastaveni seedu pro reprodukovatelnost
    if seed is not None:
        random.seed(seed)

    # Filtrace platnych operaci
    ops = [o for o in ops if o in GEN_MAP]
    if not ops:
        raise ValueError("Zadna platna operace (+ - * /).")

    # Validace poctu sloupcu
    cols = max(1, cols)

    # Vytvoreni Excel workbooku
    wb = Workbook()
    ws = wb.active
    ws.title = "Priklady"

    # Stylovani - pouzijeme monospace font pro spravne zarovnani
    font = Font(name="Consolas", size=16)
    align_left = Alignment(horizontal="left", vertical="center")

    # Hlavicka / titulek
    start_row = 1
    if title:
        title_cell = ws.cell(row=start_row, column=1)
        title_cell.value = title
        title_cell.font = Font(name="Calibri", size=18, bold=True)
        start_row += 2

    # Vygenerovani vsech prikladu najednou
    problems = [make_problem_text(ops, max_result=max_result, max_digits=max_digits, no_zero=no_zero) for _ in range(count)]

    # Validace fill_mode
    if fill_mode not in ("down", "across"):
        fill_mode = "down"

    # Vypocet potrebneho poctu radku
    rows_needed = math.ceil(count / cols)

    # Rozdeleni prikladu do sloupcu pro zjisteni max delky v kazdem sloupci
    # Nejprve vytvorime 2D strukturu prikladu podle fill_mode
    grid = [[None for _ in range(cols)] for _ in range(rows_needed)]
    idx = 0

    if fill_mode == "down":
        # Column-major: plneni po sloupcich
        for c in range(cols):
            for r in range(rows_needed):
                if idx >= count:
                    break
                grid[r][c] = problems[idx]
                idx += 1
    else:  # across
        # Row-major: plneni po radcich
        for r in range(rows_needed):
            for c in range(cols):
                if idx >= count:
                    break
                grid[r][c] = problems[idx]
                idx += 1

    # Najdeme maximalni delku leve casti (pred "= ___") v kazdem sloupci
    max_lens = [0] * cols
    for c in range(cols):
        for r in range(rows_needed):
            if grid[r][c] is not None:
                # Ziskame cast pred "= ___"
                left_part = grid[r][c].split(" = ")[0]
                max_lens[c] = max(max_lens[c], len(left_part))

    # Zarovname priklady v kazdem sloupci podle nejdelsiho
    # Mezery budou na zacatku pro zarovnani k "="
    for c in range(cols):
        for r in range(rows_needed):
            if grid[r][c] is not None:
                parts = grid[r][c].split(" = ")
                left_part = parts[0]
                # Pridame mezery na zacatek pro zarovnani zprava
                padding = max_lens[c] - len(left_part)
                grid[r][c] = " " * padding + left_part + " = ___"

    # Zapsani zarovnanych prikladu do buniek
    for r in range(rows_needed):
        for c in range(cols):
            if grid[r][c] is not None:
                cell = ws.cell(row=start_row + r, column=1 + c)
                cell.value = grid[r][c]
                cell.font = font
                cell.alignment = align_left

    # Nastaveni jednotne sirky sloupcu (cca 200 px)
    EXCEL_WIDTH = 29
    for c in range(1, cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = EXCEL_WIDTH

    # Zvyseni radkovani pro lepsi citelnost
    ROW_HEIGHT = 24
    for r in range(start_row, start_row + rows_needed):
        ws.row_dimensions[r].height = ROW_HEIGHT

    # Nastaveni uzkych okraju (narrow margins) pro tisk
    # Hodnoty jsou v palcich
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3

    # Ulozeni souboru
    wb.save(file_name)
    return file_name


# ----------------------------
# CLI rozhrani
# ----------------------------
def parse_args():
    """
    Parsuje argumenty prikazove radky.

    Returns:
        Namespace s argumenty CLI
    """
    p = argparse.ArgumentParser(
        prog="python src/main.py",
        description=f"Generator matematickych prikladu | v{__version__}",
        epilog="Pro GUI rozhrani pouzijte: python src/main.py --gui",
        formatter_class=argparse.RawDescriptionHelpFormatter
    )

    p.add_argument(
        '--version',
        action='version',
        version=f'%(prog)s {__version__}'
    )

    p.add_argument(
        "--ops",
        type=str,
        default="+-*/",
        metavar="OPERACE",
        help="Operace k pouziti. Retezec obsahujici '+' (scitani), '-' (odcitani), '*' (nasobeni), '/' (deleni). Vychozi: '+-*/'",
    )
    p.add_argument(
        "--digits",
        type=int,
        default=2,
        metavar="CISLO",
        help="Maximalni pocet cislic v cislech prikladu. Rozsah: 1-5. Vychozi: 2 (cisla 0-99). Lze kombinovat s --max.",
    )
    p.add_argument(
        "--max",
        type=int,
        default=None,
        metavar="CISLO",
        help="Maximalni vysledek prikladu. Rozsah: 1-10000. Lze kombinovat s --digits pro presnejsi kontrolu.",
    )
    p.add_argument(
        "--count",
        type=int,
        default=90,
        metavar="CISLO",
        help="Pocet prikladu k vygenerovani. Rozsah: 1-500. Vychozi: 90",
    )
    p.add_argument(
        "--seed",
        type=int,
        default=None,
        metavar="CISLO",
        help="Seed pro reprodukovatelne generovani. Pouzijte stejne cislo pro identicky vysledek. Vychozi: nahodny",
    )
    p.add_argument(
        "--out",
        type=str,
        default="priklady.xlsx",
        metavar="SOUBOR",
        help="Nazev vystupniho .xlsx souboru. Vychozi: 'priklady.xlsx'",
    )
    p.add_argument(
        "--title",
        type=str,
        default="Matematicke priklady",
        metavar="TEXT",
        help="Titulek zobrazeny v hlavicce listu. Vychozi: 'Matematicke priklady'",
    )
    p.add_argument(
        "--cols",
        type=int,
        default=3,
        metavar="CISLO",
        help="Pocet sloupcu v rozlozeni. Rozsah: 1-10. Vychozi: 3",
    )
    p.add_argument(
        "--fill",
        type=str,
        default="down",
        choices=["down", "across"],
        metavar="REZIM",
        help="Zpusob vyplnovani: 'down' (po sloupcich shora dolu) nebo 'across' (po radcich zleva doprava). Vychozi: 'down'",
    )
    p.add_argument(
        "--no-zero",
        action="store_true",
        help="Vyloucit cislo 0 z prikladu (cisla budou pouze 1 a vyse)",
    )

    return p.parse_args()


def main():
    """Hlavni entry point pro CLI aplikaci."""
    args = parse_args()
    file_path = generate_sheet(
        ops=list(args.ops),
        count=args.count,
        file_name=args.out,
        max_result=args.max,
        max_digits=args.digits,
        seed=args.seed,
        title=args.title,
        cols=args.cols,
        fill_mode=args.fill,
        no_zero=args.no_zero,
    )
    print(f"Hotovo: {file_path}")


if __name__ == "__main__":
    main()
